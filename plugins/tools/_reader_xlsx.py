
# xlsx reading, v2. The pieces:
#   coordinate grid (data islands, real row/column numbers) / formulas in their own block
#   (homogeneous fills grouped by R1C1, i.e. OOXML
#   shared-formula semantics) / empty formula cache → LibreOffice recalculates into a
#   **new file** (the source is never modified) / rendered values
#   (a light number_format renderer; LibreOffice getString is the later upgrade) /
#   merged/styles/cond
#   aggregated at sheet level / charts parsed straight from chart*.xml inside the zip
#   (openpyxl load discards charts) / pivots contribute
#   meta only, values not expanded (re-read with the cell_range parameter) / Excel Tables
#   compress through CSV (meta+preview).
# Conventions: inline code `…` = parser-added, not file content; only deviations from the defaults are noted; headers are never guessed.
import datetime as _dt
import os
import re
import shutil
import tempfile
import xml.etree.ElementTree as _ET
import zipfile

_GAP = 2          # data-island split threshold: >=2 consecutive blank rows/columns split a block (split condition diff > _GAP)
_TABLE_FULL = 20  # an Excel Table with <=20 data rows is emitted whole, otherwise meta+preview
_PREVIEW = 20     # preview row count for a large table

_NSC = "{http://schemas.openxmlformats.org/drawingml/2006/chart}"
_NSDM = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def _x_col(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)


def _x_esc(s):
    # Backtick escaping: a cell whose text contains ` must not collide with our meta markers (inline code)
    return str(s).replace("\n", " ").replace("|", "\\|").replace("`", "\\`")


def _x_ref(r1, c1, r2, c2):
    a = f"{_x_col(c1)}{r1}"
    b = f"{_x_col(c2)}{r2}"
    return a if a == b else f"{a}:{b}"


# ---------- Rendered values (light number_format rendering; anything it cannot cover degrades honestly to the stored value + fmt) ----------

def _x_render(cell):
    """cell → display text. Dates normalised to ISO; percent / scientific / thousands /
    currency / negative-in-parens rendered locally;
    unrenderable → the stored value as-is (its number_format is visible on the sheet-level
    `number-format:` line).
    No per-cell fmt/raw annotation: formats usually run down a column, so per-cell notes
    are redundant and risk clashing with content backticks."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        if v.hour == v.minute == v.second == 0:
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, (_dt.date, _dt.time)):
        return v.isoformat()
    fmt = cell.number_format or "General"
    if isinstance(v, bool) or not isinstance(v, (int, float)) or fmt in ("General", "@"):
        return _x_esc(v)
    sec = fmt.split(";")
    f = sec[1] if (v < 0 and len(sec) > 1) else sec[0]
    neg_paren = v < 0 and "(" in f
    av = abs(v) if neg_paren else v
    try:
        if "%" in f:
            m = re.search(r"0\.(0+)%", f)
            d = len(m.group(1)) if m else 0
            out = f"{av * 100:.{d}f}%"
            return f"({out})" if neg_paren else out
        if re.search(r"[0#]\.?0*E\+?0+", f, re.I):
            m = re.search(r"\.(0+)E", f, re.I)
            d = len(m.group(1)) if m else 2
            return f"{v:.{d}E}"
        if re.search(r"[0#],(?![0#])", f):  # trailing comma = thousands scaling (e.g. #,##0,,"M") → do not render
            return _x_esc(v)
        if "#,##" in f:
            m = re.search(r"0\.(0+)", f)
            d = len(m.group(1)) if m else 0
            s = f"{av:,.{d}f}"
            if "$" in f:  # any variant containing $ ('"$"', '$', the '[$$-409]' locale form)
                s = "$" + s
            return f"({s})" if neg_paren else s
    except Exception:
        pass
    return _x_esc(v)


def _x_numfmt_lines(wsf, coords):
    """Aggregate number_format onto one sheet-level line (grouped by format + range-compressed) instead of per-cell notes."""
    fmts = {}
    for (r, c) in coords:
        f = wsf.cell(row=r, column=c).number_format or "General"
        if f not in ("General", "@"):
            fmts.setdefault(f, set()).add((r, c))
    if not fmts:
        return ("number-format", [])
    bits = [f"{f} {_x_compress(cs)}" for f, cs in sorted(fmts.items())]
    return ("number-format", bits)


# ---------- recalc: when the formula cache is empty, LibreOffice recalculates into a **new file** ----------

_X_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" \
script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def _x_recalc_copy(path):
    """Copy to a new file → LibreOffice macro calculateAll recalculates and saves that copy
    → return the copy's path.
    The source is never modified. soffice unavailable or failing returns None (the caller degrades and marks it uncached)."""
    if not shutil.which("soffice"):
        return None
    try:
        env = dict(os.environ)
        env["SAL_USE_VCLPLUGIN"] = "svp"
        macro_dir = os.path.expanduser("~/.config/libreoffice/4/user/basic/Standard")
        if not os.path.isdir(macro_dir):  # first run: initialise the profile (its default contains an empty Module1)
            subprocess.run(["soffice", "--headless", "--terminate_after_init"],
                           capture_output=True, timeout=30, env=env)
        os.makedirs(macro_dir, exist_ok=True)
        mf = os.path.join(macro_dir, "Module1.xba")
        try:
            ok = "RecalculateAndSave" in open(mf, encoding="utf-8").read()
        except OSError:
            ok = False
        if not ok:
            with open(mf, "w", encoding="utf-8") as fh:
                fh.write(_X_MACRO)
        new = os.path.join(tempfile.mkdtemp(prefix="recalc_"), os.path.basename(path))
        shutil.copy2(path, new)
        cmd = ["soffice", "--headless", "--norestore",
               "vnd.sun.star.script:Standard.Module1.RecalculateAndSave"
               "?language=Basic&location=application", os.path.abspath(new)]
        r = subprocess.run(cmd, capture_output=True, timeout=90, env=env)
        return new if r.returncode == 0 else None
    except Exception:
        return None


# ---------- Data islands / range compression ----------

def _x_islands(coords):
    """Non-empty coordinate set → list of data-island bboxes. A row/column gap > _GAP
    splits a block (three passes: rows → columns → rows).
    Known risk: a title separated from its data by more than _GAP blank rows gets split off; the conservative threshold limits how often that bites."""
    if not coords:
        return []

    def split(cs, ax):
        vals = sorted({t[ax] for t in cs})
        groups, cur = [], [vals[0]]
        for v in vals[1:]:
            if v - cur[-1] > _GAP:
                groups.append(set(cur))
                cur = [v]
            else:
                cur.append(v)
        groups.append(set(cur))
        return [{t for t in cs if t[ax] in g} for g in groups]

    blocks = [set(coords)]
    for ax in (0, 1, 0):
        blocks = [b for blk in blocks for b in split(blk, ax)]
    out = []
    for b in blocks:
        rs = [t[0] for t in b]
        cs = [t[1] for t in b]
        out.append((min(rs), min(cs), max(rs), max(cs)))
    return sorted(out)


def _x_compress(coords):
    """Coordinate set → 'A1:B3,C5' rectangle compression (greedy: extend right, then down)."""
    cs = set(coords)
    out = []
    while cs:
        r, c = min(cs)
        w = 1
        while (r, c + w) in cs:
            w += 1
        h = 1
        while all((r + h, cc) in cs for cc in range(c, c + w)):
            h += 1
        for rr in range(r, r + h):
            for cc in range(c, c + w):
                cs.discard((rr, cc))
        out.append(_x_ref(r, c, r + h - 1, c + w - 1))
    return ",".join(out)


# ---------- Grid (spreadsheet view: column letters on the first row, row numbers in the first column, real coordinates) ----------

def _x_has_formula(fc):
    v = fc.value
    return (isinstance(v, str) and v.startswith("=")) or \
        v.__class__.__name__ == "ArrayFormula"


def _x_grid(wsv, wsf, bbox, mark_uncached):
    r1, c1, r2, c2 = bbox
    lines = ["|   | " + " | ".join(_x_col(c) for c in range(c1, c2 + 1)) + " |",
             "| --- |" + " --- |" * (c2 - c1 + 1)]
    for r in range(r1, r2 + 1):
        row, any_ = [], False
        for c in range(c1, c2 + 1):
            vc = wsv.cell(row=r, column=c)
            if vc.value is None and mark_uncached and _x_has_formula(wsf.cell(row=r, column=c)):
                row.append("`uncached`")
                any_ = True
                continue
            disp = _x_render(vc)
            if disp:
                any_ = True
            row.append(disp)
        if any_:
            lines.append(f"| {r} | " + " | ".join(row) + " |")
    return lines


# ---------- Formula blocks (homogeneous fill → range + R1C1; the general Excel definition, same as OOXML shared formula) ----------

_X_REF_RE = re.compile(r"(?<![A-Za-z0-9_$])(\$?)([A-Za-z]{1,3})(\$?)([0-9]{1,7})(?![\w(])")


def _x_r1c1(formula, ar, ac):
    """A1 → R1C1 (relative to anchor (ar,ac)). A homogeneous fill converts to one identical string, so it can be grouped by range."""
    from openpyxl.utils import column_index_from_string

    def conv(m):
        cd, cl, rd, rn = m.groups()
        col = column_index_from_string(cl.upper())
        row = int(rn)
        rp = f"R{row}" if rd else ("R" if row == ar else f"R[{row - ar}]")
        cp = f"C{col}" if cd else ("C" if col == ac else f"C[{col - ac}]")
        return rp + cp

    return _X_REF_RE.sub(conv, formula)


def _x_formula_lines(wsf):
    fcells = {}
    for row in wsf.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                fcells[(c.row, c.column)] = v
            elif v.__class__.__name__ == "ArrayFormula":
                t = getattr(v, "text", "") or ""
                fcells[(c.row, c.column)] = t if t.startswith("=") else "=" + t
    if not fcells:
        return ("formulas", [])
    done, items = set(), []
    for (r, c) in sorted(fcells, key=lambda t: (t[1], t[0])):  # column-major → find vertical fills first
        if (r, c) in done:
            continue
        base = _x_r1c1(fcells[(r, c)], r, c)
        k = 1
        while (r + k, c) in fcells and (r + k, c) not in done and \
                _x_r1c1(fcells[(r + k, c)], r + k, c) == base:
            k += 1
        if k > 1:
            done.update((r + i, c) for i in range(k))
            items.append(f"{_x_col(c)}{r}:{_x_col(c)}{r + k - 1} {base}")
            continue
        k = 1
        while (r, c + k) in fcells and (r, c + k) not in done and \
                _x_r1c1(fcells[(r, c + k)], r, c + k) == base:
            k += 1
        done.update((r, c + i) for i in range(max(k, 1)))
        if k > 1:
            items.append(f"{_x_col(c)}{r}:{_x_col(c + k - 1)}{r} {base}")
        else:
            items.append(f"{_x_col(c)}{r} {fcells[(r, c)]}")  # an isolated formula keeps its A1 text
    return ("formulas", items)  # one entry per formula group; the assembler puts each on its own line inside the ```meta block


# ---------- Sheet-level aggregation: merged / styles / cond / extras (links · comments · dropdowns) ----------

def _x_merged_lines(wsf):
    """For each merged region, note the cell holding the visible text (an xlsx merge always stores the value in the top-left cell)."""
    try:
        rs = []
        for r in sorted(str(r) for r in wsf.merged_cells.ranges):
            origin = r.split(":")[0]
            rs.append(f"{r} (value at {origin})")
    except Exception:
        rs = []
    return ("merged", rs)


def _x_color(c):
    try:
        if c is not None and getattr(c, "type", None) == "rgb" and c.rgb:
            s = str(c.rgb)
            if len(s) == 8:
                s = s[2:]
            if s != "000000":
                return "#" + s
    except Exception:
        pass
    return ""


def _x_style_lines(wsf, coords):
    """Styles get one line per fine-grained category: bg-color / font-color / bold / italic / underline."""
    bg, fontc = {}, {}
    bold, italic, underline = set(), set(), set()
    for (r, c) in coords:
        cell = wsf.cell(row=r, column=c)
        try:
            fill = cell.fill
            if fill is not None and fill.patternType == "solid":
                col = _x_color(fill.fgColor)
                if col and col != "#FFFFFF":
                    bg.setdefault(col, set()).add((r, c))
        except Exception:
            pass
        try:
            col = _x_color(cell.font.color)
            if col:
                fontc.setdefault(col, set()).add((r, c))
            if cell.font.bold:
                bold.add((r, c))
            if cell.font.italic:
                italic.add((r, c))
            if cell.font.underline and cell.font.underline != "none":
                underline.add((r, c))
        except Exception:
            pass
    items = []
    if bg:
        items.append("bg-color: " + " | ".join(
            f"{col} {_x_compress(cs)}" for col, cs in sorted(bg.items())))
    if fontc:
        items.append("font-color: " + " | ".join(
            f"{col} {_x_compress(cs)}" for col, cs in sorted(fontc.items())))
    if bold:
        items.append(f"bold: {_x_compress(bold)}")
    if italic:
        items.append(f"italic: {_x_compress(italic)}")
    if underline:
        items.append(f"underline: {_x_compress(underline)}")
    return ("styles", items)


def _x_cond_lines(wsf):
    bits = []
    try:
        for cf in wsf.conditional_formatting:
            rng = str(cf.sqref).replace(" ", ",")
            for rule in cf.rules:
                t = rule.type
                if t == "colorScale":
                    try:
                        # val is meaningless for min/max types → show the type name; numeric types show val
                        vs = [v.type if v.type in ("min", "max") else
                              f"{v.type}:{v.val}" for v in rule.colorScale.cfvo]
                        bits.append(f"{rng} color-scale({'→'.join(vs)})")
                    except Exception:
                        bits.append(f"{rng} color-scale")
                elif t == "dataBar":
                    bits.append(f"{rng} data-bar")
                elif t == "cellIs":
                    fml = rule.formula[0] if rule.formula else ""
                    bits.append(f"{rng} cellIs {rule.operator} {fml}")
                else:
                    bits.append(f"{rng} {t}")
    except Exception:
        pass
    return ("cond", bits)


def _x_extra_lines(wsf):
    """The remaining handles whose meaning is not in the value layer: hyperlink URLs, comments, data-validation dropdowns."""
    bits = []
    for row in wsf.iter_rows():
        for c in row:
            try:
                if c.hyperlink is not None and getattr(c.hyperlink, "target", None):
                    bits.append(f"link {c.coordinate}→{c.hyperlink.target}")
                if c.comment is not None and c.comment.text:
                    txt = c.comment.text.strip().replace("\n", " ")[:60]
                    bits.append(f'comment {c.coordinate}:"{txt}"')
            except Exception:
                pass
    try:
        for dv in wsf.data_validations.dataValidation:
            if dv.type == "list":
                bits.append(f"dropdown {dv.sqref} {dv.formula1}")
    except Exception:
        pass
    return ("extras", bits)


# ---------- Relational table → the "table with meta" representation (shared by Excel Tables and csv) ----------

def _x_to_num(v):
    """v → float or None (for the column schema's numeric verdict). bool is excluded; numeric strings (csv) count."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v)
        except ValueError:
            return None
    return None


def _x_table_md(header, rows, name=None, ref=None, preview=_PREVIEW):
    """Relational table → the "table with meta" representation (**no coordinates, no grid**),
    shared by Excel Tables and csv.
    preview = when the row count exceeds _TABLE_FULL, only the first ``preview`` rows are
    listed (Excel Tables default to _PREVIEW=20; use cell_range to fetch more);
    preview=None → everything, untruncated (used by csv, which has no cell_range; the
    layer above paginates all rows via offset/max_chars).
      ```meta
      ▸ table ["name": ref] (N data rows)
      ▸ columns                              ← only when row count > _TABLE_FULL
          col: num, min=X, max=Y             ← numeric column
          col: str, K uniq: [..]             ← text column (<=20 uniq: all listed)
          col: str, K uniq, first 20: [..]   ← text column (>20 uniq: first 20 only)
      ▸ preview: first _PREVIEW rows
      ```
      | col | col | ... |   ← markdown data table (whole thing when <=_TABLE_FULL, else the first _PREVIEW rows)
    rows: list of data rows, each cell = (raw_value, display_str). raw feeds the column schema's numeric verdict, display goes into the table body."""
    ncol = len(header)
    nrows = len(rows)
    big = nrows > _TABLE_FULL          # only a large table gets a column schema (csv gets one at full size too, as an overview)
    truncate = preview is not None and big   # Excel Tables only: a large one is cut to preview rows
    if name and ref:
        title = f'▸ table "{name}": {ref} ({nrows} data rows)'
    elif name:
        title = f'▸ table "{name}" ({nrows} data rows)'
    else:
        title = f"▸ table ({nrows} data rows)"
    lines = ["```meta", title]
    if big:
        lines.append("▸ columns")
        for ci in range(ncol):
            raws = [rows[r][ci][0] for r in range(nrows)
                    if ci < len(rows[r]) and rows[r][ci][0] not in (None, "")]
            nums = [n for n in (_x_to_num(v) for v in raws) if n is not None]
            if raws and len(nums) >= len(raws) * 0.9:
                lines.append(f"    {header[ci]}: num, min={min(nums):g}, max={max(nums):g}")
            else:
                uniq = sorted({str(v) for v in raws})
                if len(uniq) > 20:
                    lines.append(f"    {header[ci]}: str, {len(uniq)} uniq, first 20: {uniq[:20]}")
                else:
                    lines.append(f"    {header[ci]}: str, {len(uniq)} uniq: {uniq}")
        if truncate:
            lines.append(f"▸ preview: first {preview} rows")
    lines.append("```")
    lines.append("| " + " | ".join(_x_esc(h) for h in header) + " |")
    lines.append("|" + " --- |" * ncol)
    end = min(preview, nrows) if truncate else nrows
    for r in range(end):
        cells = [rows[r][ci][1] if ci < len(rows[r]) else "" for ci in range(ncol)]
        lines.append("| " + " | ".join(cells) + " |")
    return lines


def _x_table_regions(wsv, wsf):
    """Excel Table → [(top-left anchor, lines)]; inserted into the sheet's region sequence by position. Rendering goes through the shared _x_table_md."""
    from openpyxl.utils import range_boundaries
    regions, mask = [], set()
    tables = getattr(wsf, "tables", {}) or {}
    for name in tables:
        ref = tables[name].ref if hasattr(tables[name], "ref") else str(tables[name])
        try:
            c1, r1, c2, r2 = range_boundaries(ref)
        except Exception:
            continue
        mask.update((r, c) for r in range(r1, r2 + 1) for c in range(c1, c2 + 1))
        header = [str(wsv.cell(row=r1, column=c).value or "") for c in range(c1, c2 + 1)]
        rows = []
        for r in range(r1 + 1, r2 + 1):
            cells = [wsv.cell(row=r, column=c) for c in range(c1, c2 + 1)]
            rows.append([(cell.value, _x_render(cell)) for cell in cells])
        lines = _x_table_md(header, rows, name=name, ref=ref)
        regions.append(((r1, c1), lines))
    return regions, mask


def _csv_to_md(path):
    """csv/tsv → the "table with meta" representation (reuses _x_table_md).
    First row = header, the rest = data rows; the delimiter follows the extension (.tsv → tab). An empty string = an empty value.
    Output volume is controlled by _x_table_md's preview plus the layer above's max_chars/offset pagination, same as xlsx."""
    import csv as _csv
    import io as _io
    import os as _os
    sep = "\t" if path.lower().endswith(".tsv") else ","
    # Encoding via _decode_bytes (the core fragment, CJK-friendly) — hardcoding utf-8+replace
    # turns a Shift-JIS / GBK csv entirely into replacement characters
    # instead of readable text.
    with open(path, "rb") as f:
        _text = _decode_bytes(f.read())
    raw_rows = list(_csv.reader(_io.StringIO(_text, newline=""), delimiter=sep))
    name = _os.path.basename(path)
    if not raw_rows:
        return f"# CSV: {name}\n\n(empty)"
    header = [str(h) for h in raw_rows[0]]
    data = [[((v if v != "" else None), _x_esc(v)) for v in r] for r in raw_rows[1:]]
    out = ["<!-- csv readout: relational table + column meta "
           "(parser-added ```meta / `…`). -->", ""]
    out += _x_table_md(header, data, name=name, preview=None)  # csv: all rows, paginated via offset/max_chars
    return "\n".join(out)


# ---------- pivot: meta only, values not expanded (re-read with cell_range) ----------

def _x_pivot_regions(wsf):
    """pivot → [(top-left anchor, description string)]; meta only, values not expanded (re-read with cell_range)."""
    from openpyxl.utils import range_boundaries
    regions, mask = [], set()
    for p in getattr(wsf, "_pivots", []) or []:
        loc, anchor = "?", (1, 1)
        try:
            loc = p.location.ref
            c1, r1, c2, r2 = range_boundaries(loc)
            anchor = (r1, c1)
            mask.update((r, c) for r in range(r1, r2 + 1) for c in range(c1, c2 + 1))
        except Exception:
            pass
        src = "?"
        try:
            w = p.cache.cacheSource.worksheetSource
            src = f"{w.sheet}!{w.ref}"
        except Exception:
            pass
        rows, cols, vals = [], [], []
        try:
            names = [f.name for f in p.cache.cacheFields]
            rows = [names[f.x] for f in (p.rowFields or []) if 0 <= f.x < len(names)]
            cols = [names[f.x] for f in (p.colFields or []) if 0 <= f.x < len(names)]
            vals = [d.name for d in (p.dataFields or [])]
        except Exception:
            pass
        regions.append((anchor, f"at {loc} source={src} rows={rows} cols={cols} "
                                f"values={vals} — values not expanded; "
                                f"re-read with cell_range"))
    return regions, mask


# ---------- chart: parsed straight from chart*.xml inside the zip (openpyxl load discards charts, so they cannot be read back) ----------

def _x_chart_lines(path):
    out = {}
    try:
        z = zipfile.ZipFile(path)
    except Exception:
        return out
    for name in z.namelist():
        if not re.fullmatch(r"xl/charts/chart\d+\.xml", name):
            continue
        try:
            root = _ET.fromstring(z.read(name))
        except Exception:
            continue
        plot = root.find(f".//{_NSC}plotArea")
        if plot is None:
            continue
        kinds = [e.tag[len(_NSC):].replace("Chart", "") for e in plot
                 if e.tag.startswith(_NSC) and e.tag.endswith("Chart")]
        tl = root.find(f".//{_NSC}title")
        title = "".join(t.text or "" for t in tl.iter(f"{_NSDM}t")) if tl is not None else ""
        sers, cats, trends, cats_kw = [], "", [], "cats"
        for ser in plot.iter(f"{_NSC}ser"):
            v = ser.find(f"{_NSC}val")
            if v is None:
                v = ser.find(f"{_NSC}yVal")  # scatter / bubble: yVal stands in for val
            if v is not None:
                fel = v.find(f".//{_NSC}f")
                if fel is not None and fel.text:
                    sers.append(fel.text)
            ce = ser.find(f"{_NSC}cat")
            if ce is None:
                ce = ser.find(f"{_NSC}xVal")  # scatter: xVal holds x values, not categories
                if ce is not None:
                    cats_kw = "x"
            if ce is not None and not cats:
                fel = ce.find(f".//{_NSC}f")
                if fel is not None and fel.text:
                    cats = fel.text
            for t in ser.iter(f"{_NSC}trendline"):
                tt = t.find(f"{_NSC}trendlineType")
                trends.append(tt.get("val") if tt is not None else "linear")
        sheet = ""
        for ref in sers + ([cats] if cats else []):
            m = re.match(r"'?([^'!]+)'?!", ref)
            if m:
                sheet = m.group(1)
                break
        bit = "+".join(kinds) or "?"
        if title:
            bit += f' title="{title}"'
        if sers:
            bit += f" series={','.join(sers)}"
        if cats:
            bit += f" {cats_kw}={cats}"
        if trends:
            bit += f" trendline={','.join(trends)}"
        out.setdefault(sheet, []).append(bit)  # description string; the `chart N:` prefix is added once at the exit
    z.close()
    return out


# ---------- Main entry point ----------

def _xlsx_to_md(path, cell_range=None):
    """xlsx → markdown (the v2 rich extraction).
    With cell_range='Sheet1!A3:D15' only that range is read (re-reading a pivot result or part of a large file)."""
    _ensure("openpyxl", "openpyxl")
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
    # Skip eager pivotCache parsing: we never read its contents (pivot meta comes from the
    # pivotTable definition),
    # and a malformed or huge pivotCache can stall a full load_workbook (measured on one
    # dashboard: 120s → 0.09s).
    # Any cacheId → None; pivotTable definitions still parse. try/except is the floor (if
    # the openpyxl API changes, behaviour reverts to the original).
    try:
        import collections as _c

        from openpyxl.reader.workbook import WorkbookParser as _WP
        _WP.pivot_caches = property(lambda self: _c.defaultdict(lambda: None))
    except Exception:
        pass
    wbf = load_workbook(path, data_only=False)
    wbv = load_workbook(path, data_only=True)
    head = ["<!-- xlsx readout. Batch sheet metadata is collected per sheet in a ```meta"
            " fenced block (parser-added, not file content); each category is marked with"
            " a ▸ header (▸ formulas / ▸ number-format / ▸ merged / ▸ styles / ▸ cond /"
            " ▸ extras) with one item per line for quick scanning. Inline `…` marks"
            " per-cell notes (uncached / pivot / data region). Grid header = column"
            " letters, first column = row numbers (real coordinates); homogeneous formula"
            " fills grouped as R1C1; dates ISO-normalized; uncached = empty formula cache."
            " Hidden/very-hidden sheets are flagged and NOT expanded — read them on"
            " demand by passing cell_range=\"<sheet>!A1:..\". -->"]

    def needs_recalc():
        for wsf_ in wbf.worksheets:
            if wsf_.sheet_state != "visible":  # a hidden sheet is not shown and does not trigger a recalc
                continue
            wsv_ = wbv[wsf_.title]
            for row in wsf_.iter_rows():
                for c in row:
                    if _x_has_formula(c) and \
                            wsv_.cell(row=c.row, column=c.column).value is None:
                        return True
        return False

    mark_uncached = False
    if needs_recalc():
        new = _x_recalc_copy(path)
        if new:
            wbv = load_workbook(new, data_only=True)
            head.append("`recalc: done via LibreOffice on a temp copy (source untouched)`")
        else:
            mark_uncached = True
            head.append("`recalc: soffice unavailable; empty formula caches marked uncached`")

    if cell_range:  # re-read: dump only the requested range
        if "!" in cell_range:
            sn, rng = cell_range.split("!", 1)
            sn = sn.strip("'")
        else:
            sn, rng = wbv.sheetnames[0], cell_range
        if sn not in wbv.sheetnames:
            return f"[read_file error] sheet not found: {sn}"
        try:
            c1, r1, c2, r2 = range_boundaries(rng)
        except Exception as e:
            return f"[read_file error] invalid cell_range: {rng} ({e})"
        out = [*head, "", f"<!-- range {sn}!{rng} -->", ""]
        out += _x_grid(wbv[sn], wbf[sn], (r1, c1, r2, c2), mark_uncached)
        return "\n".join(out)

    charts = _x_chart_lines(path)
    out = list(head)
    for wsf in wbf.worksheets:
        if wsf.sheet_state != "visible":  # hidden sheet: annotate only, do not read the content; the LLM re-reads on demand
            state = "very hidden" if wsf.sheet_state == "veryHidden" else "hidden"
            out += ["", f"<!-- sheet: {wsf.title} -->",
                    f'`{state} — content not read; pass '
                    f'cell_range="{wsf.title}!A1:.." to read on demand`']
            continue
        wsv = wbv[wsf.title]
        out += ["", f"<!-- sheet: {wsf.title} -->"]
        coords = {(c.row, c.column) for row in wsf.iter_rows() for c in row
                  if c.value is not None}
        # Data first: data islands / Excel Tables / pivots are all emitted sorted by top-left anchor (top-left → bottom-right)
        tregions, tmask = _x_table_regions(wsv, wsf)
        pregions, pmask = _x_pivot_regions(wsf)
        islands = _x_islands(coords - tmask - pmask)
        regions = [((b[0], b[1]), "region", b) for b in islands]
        regions += [(a, "table", ls) for a, ls in tregions]
        regions += [(a, "pivot", d) for a, d in pregions]
        regions.sort(key=lambda t: t[0])
        n_reg, n_piv = len(islands), len(pregions)
        ri = 0
        pivot_items = []
        for _a, kind, payload in regions:
            if kind == "region":
                ri += 1
                label = f"data region {ri}" if n_reg > 1 else "data region"
                out += ["", f"`{label}: {_x_ref(*payload)}`"]
                out += _x_grid(wsv, wsf, payload, mark_uncached)
            elif kind == "table":
                out += ["", *payload]
            else:  # pivot: collected and placed last under ▸ pivots (values not expanded, no locating table)
                pivot_items.append(payload)
        _ = n_piv
        # Bulk sheet-level meta goes into one ```meta fence: one ▸ marker per category, entries one per line.
        # pivots/charts join it too (data first, annotations after); single-point markers (data region / uncached / recalc) stay inline.
        sections = [_x_formula_lines(wsf), _x_numfmt_lines(wsf, coords),
                    _x_merged_lines(wsf), _x_style_lines(wsf, coords),
                    _x_cond_lines(wsf), _x_extra_lines(wsf),
                    ("pivots", pivot_items),
                    ("charts", charts.get(wsf.title, []))]
        metalines = []
        for label, items in sections:
            if items:
                metalines.append(f"▸ {label}")
                metalines += [f"    {it}" for it in items]
        if metalines:
            out += ["", "```meta", *metalines, "```"]
        if not coords:
            out.append("(empty)")
    if charts.get(""):
        out += ["", "<!-- charts (sheet unresolved) -->", "```meta", "▸ charts",
                *[f"    {c}" for c in charts[""]], "```"]
    return "\n".join(out)
