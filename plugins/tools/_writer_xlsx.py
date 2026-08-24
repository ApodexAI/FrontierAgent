
# xlsx writing: incremental load-modify-save; the anchor is a sheet name + an A1 cell/range (stable, does not drift on insertion).
# Implemented with openpyxl; formulas are written as strings (=...) and computed values are the read side's job (read_file's LibreOffice recalc).
# Feature set informed by archipelago's (Apache-2.0) sheets_server; addressing and parameters are this project's own design.
#
# Operations (op):
#   create(sheets=[{name, headers?, rows?}])
#   set_cell(sheet, cell, value)               a value starting with = is treated as a formula
#   set_range(sheet, start_cell, rows=[[...]])  lay out a 2-D array starting at start_cell
#   add_sheet(sheet, headers?, rows?)
#   delete_sheet(sheet)
#   format_cells(sheet, cell_range, bold?, italic?, fill_color?, font_color?,
#                number_format?, align?)        align: left/center/right
#   add_chart(sheet, data_range, chart_type=bar|line|pie, anchor_cell="E2",
#             title?, categories_col?)
#   merge_cells(sheet, cell_range)              merge a range
#   unmerge_cells(sheet, cell_range)
#   freeze_panes(sheet, cell="B2")              freeze panes (everything above-left of cell is frozen)
#   set_column_width(sheet, columns={"A":18,...})   column width (in characters)
#   set_row_height(sheet, rows={"1":24,...})        row height (in points)
#   add_named_range(name, sheet, cell_range)    define a named range (workbook level)
#   delete_named_range(name)
#   add_data_validation(sheet, cell_range, kind=list|whole|decimal|date,
#         formula1, formula2?, prompt?, error?, allow_blank=True)   list: formula1 = a comma-separated string
#   add_conditional_formatting(sheet, cell_range, rule_type=cell_is|color_scale|
#         data_bar|formula, ...)                conditional formatting
#   set_auto_filter(sheet, cell_range)          auto filter
#   set_number_format(sheet, cell_range, number_format)
#   add_image(sheet, image_path, anchor_cell="A1")
import os as _os


def _xl_hex(c):
    """Normalise a 6- or 8-digit hex colour to openpyxl's ARGB (8 digits, with an FF alpha)."""
    h = str(c).lstrip("#").upper()
    return h if len(h) == 8 else "FF" + h


def _xl():
    _ensure("openpyxl", "openpyxl")
    import openpyxl
    return openpyxl


def _coerce(v):
    """Convert a numeric string to int/float; a leading = keeps it as a formula string."""
    if not isinstance(v, str):
        return v
    if v.startswith("="):
        return v
    try:
        return int(v)
    except ValueError:
        try:
            return float(v)
        except ValueError:
            return v


# Named number_format enum (raw Excel format strings are accepted too)
_NUMFMT = {
    "general": "General", "integer": "0", "number2": "0.00",
    "percent": "0%", "percent2": "0.00%",
    "currency_usd": '"$"#,##0.00', "currency_eur": '"€"#,##0.00',
    "accounting": '_("$"* #,##0.00_)', "date_iso": "yyyy-mm-dd",
    "date_us": "m/d/yyyy", "datetime": "yyyy-mm-dd hh:mm", "time": "hh:mm:ss",
    "scientific": "0.00E+00", "text": "@",
}


def _numfmt(v):
    return _NUMFMT.get(str(v), v) if v else v


def _typed_value(value, typ):
    """Write a value under an explicit type (cures "numbers stored as text"). type: auto|number|text|formula|date|bool."""
    import datetime as _dt
    t = (typ or "auto").lower()
    if t == "auto":
        return _coerce(value)
    if t == "text":
        return "" if value is None else str(value)
    if t == "number":
        try:
            return int(value)
        except (ValueError, TypeError):
            try:
                return float(value)
            except (ValueError, TypeError):
                return value
    if t == "formula":
        s = str(value)
        return s if s.startswith("=") else "=" + s
    if t == "bool":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ("true", "1", "yes", "y")
    if t == "date":
        if isinstance(value, (_dt.date, _dt.datetime)):
            return value
        s = str(value)
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"):
            try:
                return _dt.datetime.strptime(s, fmt)
            except ValueError:
                continue
        return value
    return _coerce(value)


def _autofit_columns(ws):
    """Estimate column width from content (deterministic, no LLM): the longest content per column + 2, clamped to [8, 60]."""
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and getattr(c, "column_letter", None):
                widths[c.column_letter] = max(widths.get(c.column_letter, 0), len(str(c.value)))
    for col, w in widths.items():
        ws.column_dimensions[col].width = min(max(w + 2, 8), 60)


def _save_ret(wb, path, msg):
    """Save + return the message (removes the wb.save boilerplate at the end of every op)."""
    wb.save(path)
    return msg


def _get_ws(wb, sheet):
    if sheet is None:
        return wb.active
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet!r} (have {wb.sheetnames})")
    return wb[sheet]


def _fill_sheet(ws, headers, rows):
    """Whether a formula was written (fixes: create/add_sheet carrying formulas did not set wrote_formula,
    so the recalc gate never fired and formula caches stayed empty)."""
    wrote_formula = False
    if headers:
        ws.append(list(headers))
        ws.freeze_panes = "A2"
    for row in (rows or []):
        vals = [_coerce(x) for x in row]
        if not wrote_formula and any(isinstance(v, str) and v.startswith("=") for v in vals):
            wrote_formula = True
        ws.append(vals)
    _autofit_columns(ws)  # default to content-based column widths (deterministic, overridable)
    return wrote_formula


def _xlsx_write(path, op, args):
    openpyxl = _xl()

    if op == "create":
        if _os.path.exists(path) and not args.get("overwrite"):
            return _res(f"create refused: {path} already exists — use set_cell/add_sheet to "
                        "edit, or pass overwrite:true to rebuild", ok=False)
        wb = openpyxl.Workbook()
        sheets = args.get("sheets") or [{"name": "Sheet1"}]
        wb.remove(wb.active)
        wf = False
        for sd in sheets:
            ws = wb.create_sheet(title=(sd.get("name") or "Sheet")[:31])
            wf = _fill_sheet(ws, sd.get("headers"), sd.get("rows")) or wf
        _os.makedirs(_os.path.dirname(path) or ".", exist_ok=True)
        wb.save(path)
        return _res(f"created xlsx: {path}", counts={"sheet": len(sheets)}, wrote_formula=wf)

    if not _os.path.exists(path):
        return f"[error] file not found (edit needs existing file): {path}"
    wb = openpyxl.load_workbook(path)

    if op == "set_cell":
        ws = _get_ws(wb, args.get("sheet"))
        cell = ws[args["cell"]]
        tv = _typed_value(args.get("value"), args.get("type"))
        cell.value = tv
        if args.get("number_format"):
            cell.number_format = _numfmt(args["number_format"])
        wb.save(path)
        return _res(f"set {args.get('sheet') or ws.title}!{args['cell']} = {args.get('value')!r}",
                    wrote_formula=isinstance(tv, str) and tv.startswith("="))

    if op == "set_range":
        ws = _get_ws(wb, args.get("sheet"))
        from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
        r0, c0 = coordinate_to_tuple(args["start_cell"])
        types = args.get("types")
        n, wf = 0, False
        for i, row in enumerate(args["rows"]):
            for j, val in enumerate(row):
                if isinstance(types, list):
                    tt = types[i][j] if (i < len(types) and isinstance(types[i], list)
                                         and j < len(types[i])) else None
                else:
                    tt = types
                tv = _typed_value(val, tt)
                ws[f"{get_column_letter(c0 + j)}{r0 + i}"] = tv
                wf = wf or (isinstance(tv, str) and tv.startswith("="))
                n += 1
        wb.save(path)
        return _res(f"set {n} cell(s) from {args['start_cell']}", wrote_formula=wf)

    if op == "add_sheet":
        name = (args["sheet"])[:31]
        if name in wb.sheetnames:
            return f"[error] sheet already exists: {name}"
        ws = wb.create_sheet(title=name)
        wf = _fill_sheet(ws, args.get("headers"), args.get("rows"))
        wb.save(path)
        return _res(f"added sheet {name!r}", wrote_formula=wf)

    if op == "delete_sheet":
        ws = _get_ws(wb, args.get("sheet"))
        if len(wb.sheetnames) == 1:
            return "[error] cannot delete the only sheet"
        wb.remove(ws)
        wb.save(path)
        return f"deleted sheet {args.get('sheet')!r}"

    if op in ("format_cells", "set_cell_format"):
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        ws = _get_ws(wb, args.get("sheet"))
        # Normalise into a flat cell list (a single-cell range is simply wrapped; the previous
        # subscript access on a single cell was dead code that raised TypeError and aborted the whole batch, so it is gone)
        flat = []
        rng = ws[args["cell_range"]]
        if hasattr(rng, "value"):           # single cell
            flat = [rng]
        else:
            for row in rng:
                flat.extend(row if isinstance(row, tuple) else [row])
        fkw = {}
        if args.get("bold") is not None:
            fkw["bold"] = bool(args["bold"])
        if args.get("italic") is not None:
            fkw["italic"] = bool(args["italic"])
        if args.get("font_color"):
            fkw["color"] = str(args["font_color"]).lstrip("#")
        if args.get("font_size"):
            fkw["size"] = float(args["font_size"])
        if args.get("font_name"):
            fkw["name"] = args["font_name"]
        # alignment(align_h / align_v / wrap)
        akw = {}
        ah = args.get("align_h") or args.get("align")
        if ah:
            akw["horizontal"] = str(ah).lower()
        if args.get("align_v"):
            akw["vertical"] = {"middle": "center"}.get(str(args["align_v"]).lower(),
                                                       str(args["align_v"]).lower())
        if args.get("wrap") is not None:
            akw["wrap_text"] = bool(args["wrap"])
        # border
        border = None
        if args.get("border"):
            b = args["border"]
            style = str(b.get("style", "thin")).lower()
            col = _xl_hex(b["color"]) if b.get("color") else None
            side = Side(style=(None if style == "none" else style), color=col)
            sides = b.get("sides", "all")
            sides = [sides] if isinstance(sides, str) else list(sides)
            want = set()
            for s in sides:
                s = str(s).lower()
                if s in ("all", "outline"):
                    want |= {"top", "bottom", "left", "right"}
                else:
                    want.add(s)
            border = Border(**{k: side for k in want})
        from copy import copy as _cp

        from openpyxl.styles import Color as _Color
        for c in flat:
            if fkw:
                # Incremental font edit: copy the existing Font and override attribute by
                # attribute (fixes: Font(**fkw) replaces wholesale and resets
                # untouched name/size/color/underline to defaults, breaking "leave what was not touched alone")
                nf = _cp(c.font)
                for k, v in fkw.items():
                    setattr(nf, k, _Color(rgb=str(v)) if k == "color" else v)
                c.font = nf
            if args.get("fill_color"):
                c.fill = PatternFill("solid", fgColor=str(args["fill_color"]).lstrip("#"))
            if args.get("number_format"):
                c.number_format = _numfmt(args["number_format"])
            if akw:
                c.alignment = Alignment(**akw)
            if border is not None:
                c.border = border
        wb.save(path)
        return f"formatted {len(flat)} cell(s) in {args['cell_range']}"

    if op == "add_chart":
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        ws = _get_ws(wb, args.get("sheet"))
        kind = args.get("chart_type", "bar")
        chart = {"bar": BarChart, "line": LineChart, "pie": PieChart}.get(kind, BarChart)()
        if args.get("title"):
            chart.title = args["title"]
        from openpyxl.utils.cell import range_boundaries
        c1, r1, c2, r2 = range_boundaries(args["data_range"])
        cat_col = args.get("categories_col")
        hdr = args.get("include_header", True)
        data_min_col = c1 + 1 if cat_col in (None, c1) else c1
        data = Reference(ws, min_col=(c1 + 1), min_row=r1, max_col=c2, max_row=r2)
        chart.add_data(data, titles_from_data=hdr)
        cats = Reference(ws, min_col=c1, min_row=(r1 + 1 if hdr else r1), max_row=r2)
        chart.set_categories(cats)
        _ = data_min_col
        if args.get("width"):          # chart size (centimetres, openpyxl's unit)
            chart.width = float(args["width"])
        if args.get("height"):
            chart.height = float(args["height"])
        ws.add_chart(chart, args.get("anchor_cell", "E2"))
        wb.save(path)
        return f"added {kind} chart from {args['data_range']}"

    if op == "merge_cells":
        ws = _get_ws(wb, args.get("sheet"))
        ws.merge_cells(args["cell_range"])
        wb.save(path)
        return f"merged {args['cell_range']}"

    if op == "unmerge_cells":
        ws = _get_ws(wb, args.get("sheet"))
        ws.unmerge_cells(args["cell_range"])
        wb.save(path)
        return f"unmerged {args['cell_range']}"

    if op == "freeze_panes":
        ws = _get_ws(wb, args.get("sheet"))
        ws.freeze_panes = args.get("cell") or args.get("cell_range") or "A2"
        wb.save(path)
        return f"froze panes at {ws.freeze_panes}"

    if op == "set_column_width":
        ws = _get_ws(wb, args.get("sheet"))
        cols = args.get("columns") or {}
        for col, w in cols.items():
            ws.column_dimensions[str(col).upper()].width = float(w)
        for col in (args.get("hidden") or []):     # hidden columns (list of column names)
            ws.column_dimensions[str(col).upper()].hidden = True
        return _save_ret(wb, path, f"set width on {len(cols)} column(s)"
                         + (f", hid {len(args['hidden'])}" if args.get("hidden") else ""))

    if op == "set_row_height":
        ws = _get_ws(wb, args.get("sheet"))
        rows = args.get("rows") or {}
        for r, h in rows.items():
            ws.row_dimensions[int(r)].height = float(h)
        for r in (args.get("hidden") or []):        # hidden rows (list of row numbers)
            ws.row_dimensions[int(r)].hidden = True
        return _save_ret(wb, path, f"set height on {len(rows)} row(s)"
                         + (f", hid {len(args['hidden'])}" if args.get("hidden") else ""))

    if op == "set_page_setup":
        # xlsx print / page properties (the biggest gap in evaluation: landscape / fit-to-page / margins / print titles when a sheet does not fit one page).
        ws = _get_ws(wb, args.get("sheet"))
        ps = ws.page_setup
        o = args.get("orientation")
        if o:
            ps.orientation = "landscape" if str(o).lower().startswith("land") else "portrait"
        fw, fh = args.get("fit_to_width"), args.get("fit_to_height")
        if fw is not None or fh is not None or args.get("fit_to_page"):
            from openpyxl.worksheet.properties import PageSetupProperties
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            if fw is not None:
                ps.fitToWidth = int(fw)
            if fh is not None:
                ps.fitToHeight = int(fh)
        if args.get("scale") is not None:
            ps.scale = int(args["scale"])
        if args.get("paper_size"):
            _paper = {"a4": 9, "letter": 1, "legal": 5, "a3": 8, "tabloid": 3, "a5": 11}
            v = _paper.get(str(args["paper_size"]).lower())
            if v:
                ps.paperSize = v
        m = args.get("margins") or {}
        if m:
            from openpyxl.worksheet.page import PageMargins
            ws.page_margins = PageMargins(**{k: float(val) for k, val in m.items()
                                             if k in ("left", "right", "top", "bottom",
                                                      "header", "footer")})
        if args.get("center_h") is not None:
            ws.print_options.horizontalCentered = bool(args["center_h"])
        if args.get("center_v") is not None:
            ws.print_options.verticalCentered = bool(args["center_v"])
        if args.get("print_area"):
            ws.print_area = args["print_area"]
        if args.get("print_title_rows"):      # title rows repeated on every page, e.g. "1:1"
            ws.print_title_rows = args["print_title_rows"]
        if args.get("print_title_cols"):      # title columns repeated on every page, e.g. "A:A"
            ws.print_title_cols = args["print_title_cols"]
        return _save_ret(wb, path, f"set page setup on {ws.title}")

    if op == "rename_sheet":
        ws = _get_ws(wb, args.get("sheet") or args.get("old"))
        new = args.get("new") or args.get("new_name")
        if not new:
            return "[error] rename_sheet needs 'new'"
        old = ws.title
        ws.title = str(new)[:31]
        return _save_ret(wb, path, f"renamed sheet {old!r} -> {ws.title!r}")

    if op in ("hide_sheet", "show_sheet", "set_sheet_visibility"):
        ws = _get_ws(wb, args.get("sheet"))
        hidden = True if op == "hide_sheet" else (False if op == "show_sheet"
                                                  else bool(args.get("hidden", True)))
        ws.sheet_state = "hidden" if hidden else "visible"
        if all(w.sheet_state != "visible" for w in wb.worksheets):
            ws.sheet_state = "visible"
            return "[error] cannot hide the only visible sheet"
        return _save_ret(wb, path, f"sheet {ws.title!r} -> {ws.sheet_state}")

    if op == "clear_charts":
        ws = _get_ws(wb, args.get("sheet"))
        n = len(ws._charts)
        ws._charts = []
        return _save_ret(wb, path, f"cleared {n} chart(s) on {ws.title}")

    if op == "add_named_range":
        from openpyxl.utils import quote_sheetname
        from openpyxl.workbook.defined_name import DefinedName
        ws = _get_ws(wb, args.get("sheet"))
        name = args["name"]
        ref = f"{quote_sheetname(ws.title)}!{args['cell_range']}"
        # Absolute references ($) are more robust; if the user did not supply $, use it as given
        wb.defined_names.add(DefinedName(name, attr_text=ref))
        wb.save(path)
        return f"defined name {name!r} -> {ref}"

    if op == "delete_named_range":
        name = args["name"]
        if name in wb.defined_names:
            del wb.defined_names[name]
            wb.save(path)
            return f"deleted name {name!r}"
        return f"[error] named range not found: {name!r}"

    if op == "add_data_validation":
        from openpyxl.worksheet.datavalidation import DataValidation
        ws = _get_ws(wb, args.get("sheet"))
        kind = args.get("kind", "list")
        f1 = args.get("formula1", "")
        if kind == "list" and isinstance(f1, (list, tuple)):
            f1 = ",".join(str(x) for x in f1)
        if kind == "list" and not str(f1).startswith('"') and "," in str(f1) and "!" not in str(f1):
            f1 = f'"{f1}"'  # an inline list has to be wrapped in quotes
        dv = DataValidation(
            type=kind, formula1=f1, formula2=args.get("formula2"),
            operator=args.get("operator", "between"),
            allow_blank=bool(args.get("allow_blank", True)),
            showDropDown=False,
        )
        if args.get("prompt"):
            dv.prompt = args["prompt"]
            dv.promptTitle = args.get("prompt_title", "")
        if args.get("error"):
            dv.error = args["error"]
            dv.errorTitle = args.get("error_title", "")
            dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(args["cell_range"])
        wb.save(path)
        return f"added {kind} validation on {args['cell_range']}"

    if op == "add_conditional_formatting":
        ws = _get_ws(wb, args.get("sheet"))
        rng = args["cell_range"]
        rtype = args.get("rule_type", "cell_is")
        rule = None
        if rtype == "cell_is":
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import Font, PatternFill
            kw = {}
            if args.get("fill_color"):
                kw["fill"] = PatternFill("solid", fgColor=_xl_hex(args["fill_color"]))
            if args.get("font_color"):
                kw["font"] = Font(color=_xl_hex(args["font_color"]))
            rule = CellIsRule(
                operator=args.get("operator", "greaterThan"),
                formula=[str(x) for x in (args.get("formula") or [args.get("value", 0)])],
                **kw)
        elif rtype == "color_scale":
            from openpyxl.formatting.rule import ColorScaleRule
            cols = args.get("colors") or ["FFFF0000", "FFFFFF00", "FF00FF00"]
            cols = [_xl_hex(c) for c in cols]
            if len(cols) == 2:
                rule = ColorScaleRule(start_type="min", start_color=cols[0],
                                      end_type="max", end_color=cols[1])
            else:
                rule = ColorScaleRule(start_type="min", start_color=cols[0],
                                      mid_type="percentile", mid_value=50, mid_color=cols[1],
                                      end_type="max", end_color=cols[2])
        elif rtype == "data_bar":
            from openpyxl.formatting.rule import DataBarRule
            rule = DataBarRule(start_type="min", end_type="max",
                               color=_xl_hex(args.get("color", "FF638EC6")))
        elif rtype == "formula":
            from openpyxl.formatting.rule import FormulaRule
            from openpyxl.styles import Font, PatternFill
            kw = {}
            if args.get("fill_color"):
                kw["fill"] = PatternFill("solid", fgColor=_xl_hex(args["fill_color"]))
            if args.get("font_color"):
                kw["font"] = Font(color=_xl_hex(args["font_color"]))
            rule = FormulaRule(formula=args["formula"], **kw)
        if rule is None:
            return f"[error] unknown rule_type: {rtype}"
        ws.conditional_formatting.add(rng, rule)
        wb.save(path)
        return f"added {rtype} conditional formatting on {rng}"

    if op == "set_auto_filter":
        ws = _get_ws(wb, args.get("sheet"))
        ws.auto_filter.ref = args.get("cell_range") or args.get("ref") or ws.dimensions
        wb.save(path)
        return f"set auto_filter on {ws.auto_filter.ref}"

    if op == "set_number_format":
        ws = _get_ws(wb, args.get("sheet"))
        fmt = _numfmt(args["number_format"])
        rng = ws[args["cell_range"]]
        cells = rng if isinstance(rng, tuple) else [(rng,)]
        n = 0
        for row in cells:
            for c in (row if isinstance(row, tuple) else [row]):
                c.number_format = fmt
                n += 1
        wb.save(path)
        return f"set number_format on {n} cell(s)"

    if op == "add_image":
        from openpyxl.drawing.image import Image as _XLImage
        ws = _get_ws(wb, args.get("sheet"))
        img = _XLImage(args["image_path"])
        if args.get("width"):
            img.width = int(args["width"])
        if args.get("height"):
            img.height = int(args["height"])
        ws.add_image(img, args.get("anchor_cell", "A1"))
        wb.save(path)
        return f"added image at {args.get('anchor_cell', 'A1')}"

    if op == "add_table":
        from openpyxl.worksheet.table import Table, TableStyleInfo
        ws = _get_ws(wb, args.get("sheet"))
        ref = args["cell_range"]
        name = args.get("name") or f"Table{len(ws.tables) + 1}"
        # headers is optional: when supplied and the range's first row is still empty, write the header first (a ListObject requires unique column names on the first row)
        if args.get("headers"):
            from openpyxl.utils.cell import get_column_letter, range_boundaries
            c1, r1, _c2, _r2 = range_boundaries(ref)
            for j, h in enumerate(args["headers"]):
                ws[f"{get_column_letter(c1 + j)}{r1}"] = h
        tbl = Table(displayName=name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name=args.get("style", "TableStyleMedium9"),
            showRowStripes=True, showColumnStripes=False,
            showFirstColumn=False, showLastColumn=False)
        ws.add_table(tbl)  # a ListObject brings its own autofilter
        wb.save(path)
        return f"added table {name!r} over {ref} (ListObject)"

    return f"[error] unknown xlsx op: {op}"
